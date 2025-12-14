"""File and spreadsheet tools.

Tools for file operations:
- file_writer: Write content to files
- file_reader: Read content from files
- spreadsheet_read: Read Excel spreadsheets
- spreadsheet_write: Write Excel spreadsheets
"""

import base64
import json
import logging
import os
import tempfile
import time
from typing import Any
from urllib.parse import quote

from .base import BuiltinToolResult, UserContext

logger = logging.getLogger(__name__)

# Workspace TTL in hours
WORKSPACE_FILE_TTL_HOURS = 24
# Maximum cell value length
MAX_CELL_VALUE_LENGTH = 500


def _get_workspace_dir(user_context: UserContext | None) -> str:
    """Get the agent workspace directory for the current user."""
    base_dir = os.path.join(tempfile.gettempdir(), "agent_workspace")

    if user_context:
        safe_user_id = "".join(c for c in user_context.user_id if c.isalnum() or c in "-_")
        if safe_user_id:
            workspace_dir = os.path.join(base_dir, safe_user_id)
        else:
            workspace_dir = os.path.join(base_dir, "anonymous")
    else:
        workspace_dir = os.path.join(base_dir, "anonymous")

    os.makedirs(workspace_dir, exist_ok=True)
    return workspace_dir


def _get_download_url(filename: str) -> str:
    """Generate a download URL for a workspace file."""
    return f"/api/files/{quote(filename)}"


def _cleanup_old_files(workspace_dir: str, ttl_hours: int = 24) -> int:
    """Remove files older than TTL from workspace."""
    removed = 0
    now = time.time()
    ttl_seconds = ttl_hours * 3600

    try:
        for filename in os.listdir(workspace_dir):
            filepath = os.path.join(workspace_dir, filename)
            if os.path.isfile(filepath):
                file_age = now - os.path.getmtime(filepath)
                if file_age > ttl_seconds:
                    os.remove(filepath)
                    removed += 1
                    logger.debug(f"Cleaned up old file: {filename}")
    except Exception as e:
        logger.warning(f"Workspace cleanup error: {e}")

    return removed


def _truncate_cell_value(value: Any, max_length: int | None = None) -> Any:
    """Truncate cell values that are too long."""
    if value is None:
        return None

    max_len = max_length or MAX_CELL_VALUE_LENGTH

    if isinstance(value, str) and len(value) > max_len:
        return value[:max_len] + f"... [truncated, {len(value)} chars total]"

    return value


async def execute_file_writer(arguments: dict[str, Any], user_context: UserContext | None = None) -> BuiltinToolResult:
    """Execute the file_writer tool."""
    filename = arguments.get("filename", "")
    content = arguments.get("content", "")
    mode = arguments.get("mode", "overwrite")
    is_binary = arguments.get("is_binary", False)

    if not filename:
        return BuiltinToolResult(success=False, error="Filename is required")
    if not content:
        return BuiltinToolResult(success=False, error="Content is required")

    if ".." in filename or filename.startswith("/"):
        return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

    ext = os.path.splitext(filename)[1].lower()

    text_extensions = {".txt", ".md", ".json", ".csv", ".py", ".js", ".html", ".css", ".xml", ".yaml", ".yml"}
    binary_extensions = {".xlsx", ".xls", ".pdf", ".png", ".jpg", ".jpeg", ".gif", ".zip", ".docx", ".pptx"}
    allowed_extensions = text_extensions | binary_extensions

    if ext not in allowed_extensions:
        return BuiltinToolResult(
            success=False,
            error=f"File extension '{ext}' not allowed. Allowed: {', '.join(sorted(allowed_extensions))}",
        )

    if ext in binary_extensions and not is_binary:
        return BuiltinToolResult(
            success=False,
            error=f"Binary extension '{ext}' requires is_binary=True with base64 content.",
        )

    content_size = len(content.encode("utf-8")) if not is_binary else len(content)
    if content_size > 5 * 1024 * 1024:
        return BuiltinToolResult(success=False, error="Content exceeds 5MB limit")

    logger.info(f"Writing file: {filename} ({len(content)} chars, binary={is_binary})")

    try:
        workspace_dir = _get_workspace_dir(user_context)
        _cleanup_old_files(workspace_dir)
        file_path = os.path.join(workspace_dir, filename)

        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else workspace_dir, exist_ok=True)

        if is_binary:
            try:
                binary_content = base64.b64decode(content)
            except Exception as e:
                return BuiltinToolResult(success=False, error=f"Invalid base64 content: {str(e)}")

            write_mode = "ab" if mode == "append" else "wb"
            with open(file_path, write_mode) as f:
                f.write(binary_content)

            size_bytes = len(binary_content)
        else:
            write_mode = "a" if mode == "append" else "w"
            with open(file_path, write_mode, encoding="utf-8") as f:
                f.write(content)

            size_bytes = len(content.encode("utf-8"))

        download_url = _get_download_url(filename)

        return BuiltinToolResult(
            success=True,
            result={
                "filename": filename,
                "path": file_path,
                "size_bytes": size_bytes,
                "mode": mode,
                "is_binary": is_binary,
                "download_url": download_url,
                "ttl_hours": WORKSPACE_FILE_TTL_HOURS,
                "note": f"File available for download at {download_url}. Expires in {WORKSPACE_FILE_TTL_HOURS} hours.",
            },
        )

    except Exception as e:
        logger.exception(f"File write failed: {e}")
        return BuiltinToolResult(success=False, error=f"Write failed: {str(e)}")


async def execute_file_reader(arguments: dict[str, Any], user_context: UserContext | None = None) -> BuiltinToolResult:
    """Execute the file_reader tool."""
    filename = arguments.get("filename", "")
    encoding = arguments.get("encoding", "utf-8")

    if not filename:
        return BuiltinToolResult(success=False, error="Filename is required")

    if ".." in filename or filename.startswith("/"):
        return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

    logger.info(f"Reading file: {filename}")

    try:
        workspace_dir = _get_workspace_dir(user_context)
        file_path = os.path.join(workspace_dir, filename)

        if not os.path.exists(file_path):
            return BuiltinToolResult(success=False, error=f"File not found: {filename}")

        file_size = os.path.getsize(file_path)
        _, ext = os.path.splitext(filename.lower())

        spreadsheet_extensions = {".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}
        document_extensions = {".docx", ".doc", ".pptx", ".ppt", ".odt", ".odp"}
        binary_extensions = {".pdf", ".zip", ".tar", ".gz", ".rar", ".7z", ".exe", ".dll", ".so", ".dylib"}
        image_extensions = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".webp", ".ico"}

        if ext in spreadsheet_extensions:
            return BuiltinToolResult(
                success=False,
                error=f"'{filename}' is a spreadsheet file. Use 'spreadsheet_read' tool instead.",
                metadata={"filename": filename, "size_bytes": file_size, "file_type": "spreadsheet"},
            )

        if ext in document_extensions:
            return BuiltinToolResult(
                success=False,
                error=f"'{filename}' is a binary document file ({ext}). File size: {file_size:,} bytes.",
                metadata={"filename": filename, "size_bytes": file_size, "file_type": "document"},
            )

        if ext in binary_extensions:
            return BuiltinToolResult(
                success=False,
                error=f"'{filename}' is a binary file ({ext}). File size: {file_size:,} bytes.",
                metadata={"filename": filename, "size_bytes": file_size, "file_type": "binary"},
            )

        if ext in image_extensions:
            return BuiltinToolResult(
                success=False,
                error=f"'{filename}' is an image file ({ext}). File size: {file_size:,} bytes.",
                metadata={"filename": filename, "size_bytes": file_size, "file_type": "image"},
            )

        max_text_size = 500_000

        with open(file_path, encoding=encoding) as f:
            content = f.read()

        if len(content) > max_text_size:
            content = content[:max_text_size]
            return BuiltinToolResult(
                success=True,
                result=content,
                metadata={
                    "filename": filename,
                    "size_bytes": file_size,
                    "truncated": True,
                    "note": f"Content truncated to {max_text_size:,} characters.",
                },
            )

        return BuiltinToolResult(
            success=True,
            result=content,
            metadata={"filename": filename, "size_bytes": len(content.encode("utf-8"))},
        )

    except UnicodeDecodeError:
        try:
            file_size = os.path.getsize(file_path)
            return BuiltinToolResult(
                success=False,
                error=f"'{filename}' appears to be a binary file. File size: {file_size:,} bytes.",
                metadata={"filename": filename, "size_bytes": file_size, "file_type": "binary"},
            )
        except Exception as e:
            return BuiltinToolResult(success=False, error=f"Read failed: {str(e)}")
    except Exception as e:
        logger.exception(f"File read failed: {e}")
        return BuiltinToolResult(success=False, error=f"Read failed: {str(e)}")


async def execute_spreadsheet_read(arguments: dict[str, Any], user_context: UserContext | None = None) -> BuiltinToolResult:
    """Execute the spreadsheet_read tool."""
    filename = arguments.get("filename", "")
    sheet_name = arguments.get("sheet_name")
    include_stats = arguments.get("include_stats", True)
    max_rows = min(arguments.get("max_rows", 50), 500)
    offset = arguments.get("offset", 0)
    columns_filter = arguments.get("columns")

    MAX_RESULT_SIZE = 100_000

    if not filename:
        return BuiltinToolResult(success=False, error="Filename is required")

    if ".." in filename or filename.startswith("/"):
        return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

    if not filename.lower().endswith(".xlsx"):
        return BuiltinToolResult(success=False, error="Only .xlsx files are supported")

    logger.info(f"Reading spreadsheet: {filename}")

    try:
        from openpyxl import load_workbook

        workspace_dir = _get_workspace_dir(user_context)
        _cleanup_old_files(workspace_dir)
        file_path = os.path.join(workspace_dir, filename)

        if not os.path.exists(file_path):
            return BuiltinToolResult(success=False, error=f"File not found: {filename}")

        wb = load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        all_rows = list(ws.iter_rows(values_only=True))
        total_rows = len(all_rows)

        if total_rows == 0:
            wb.close()
            return BuiltinToolResult(
                success=True,
                result={
                    "filename": filename,
                    "sheet_name": sheet_name,
                    "sheets": wb.sheetnames,
                    "total_rows": 0,
                    "headers": [],
                    "data": [],
                },
            )

        headers = [str(h) if h is not None else f"Column_{i + 1}" for i, h in enumerate(all_rows[0])]

        column_indices = None
        if columns_filter:
            column_indices = [i for i, h in enumerate(headers) if h in columns_filter]
            if not column_indices:
                wb.close()
                return BuiltinToolResult(success=False, error=f"No matching columns found. Available: {headers}")
            headers = [headers[i] for i in column_indices]

        data_rows = all_rows[1:]
        total_data_rows = len(data_rows)

        start_idx = offset
        end_idx = min(offset + max_rows, total_data_rows)
        paginated_rows = data_rows[start_idx:end_idx]

        data = []
        current_size = 0
        truncated_at_row = None

        for row_idx, row in enumerate(paginated_rows):
            if column_indices:
                row_data = {headers[j]: _truncate_cell_value(row[i]) for j, i in enumerate(column_indices)}
            else:
                row_data = {headers[i]: _truncate_cell_value(val) for i, val in enumerate(row)}

            row_json = json.dumps(row_data, default=str)
            row_size = len(row_json)

            if current_size + row_size > MAX_RESULT_SIZE:
                truncated_at_row = row_idx
                break

            data.append(row_data)
            current_size += row_size

        result = {
            "filename": filename,
            "sheet_name": sheet_name,
            "sheets": wb.sheetnames,
            "total_rows": total_data_rows,
            "returned_rows": len(data),
            "offset": offset,
            "headers": headers,
            "data": data,
        }

        if truncated_at_row is not None:
            result["truncated"] = True
            result["note"] = f"Results truncated to {len(data)} rows due to size limits. Use 'offset' parameter to paginate through remaining rows."

        if include_stats and data:
            stats = {"columns": {}}
            for header in headers:
                col_values = [row.get(header) for row in data if row.get(header) is not None]
                col_stats = {"count": len(col_values)}

                numeric_values = [v for v in col_values if isinstance(v, int | float)]

                if numeric_values:
                    col_stats["type"] = "numeric"
                    col_stats["min"] = min(numeric_values)
                    col_stats["max"] = max(numeric_values)
                    col_stats["avg"] = sum(numeric_values) / len(numeric_values)
                else:
                    col_stats["type"] = "text"
                    col_stats["unique_count"] = len(set(str(v) for v in col_values))

                stats["columns"][header] = col_stats

            result["stats"] = stats

        wb.close()

        return BuiltinToolResult(
            success=True,
            result=result,
            metadata={"filename": filename, "total_rows": total_data_rows, "returned_rows": len(data)},
        )

    except Exception as e:
        logger.exception(f"Spreadsheet read failed: {e}")
        return BuiltinToolResult(success=False, error=f"Read failed: {str(e)}")


async def execute_spreadsheet_write(arguments: dict[str, Any], user_context: UserContext | None = None) -> BuiltinToolResult:
    """Execute the spreadsheet_write tool."""
    filename = arguments.get("filename", "")
    operation = arguments.get("operation", "create")
    sheet_name = arguments.get("sheet_name", "Sheet1")
    headers = arguments.get("headers", [])
    data = arguments.get("data", [])
    cell_updates = arguments.get("cell_updates", [])

    if not filename:
        return BuiltinToolResult(success=False, error="Filename is required")

    if ".." in filename or filename.startswith("/"):
        return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

    if not filename.lower().endswith(".xlsx"):
        filename = filename + ".xlsx"

    logger.info(f"Writing spreadsheet: {filename} (operation: {operation})")

    try:
        from openpyxl import Workbook, load_workbook

        workspace_dir = _get_workspace_dir(user_context)
        _cleanup_old_files(workspace_dir)
        file_path = os.path.join(workspace_dir, filename)

        if operation == "create":
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            if headers:
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            start_row = 2 if headers else 1
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(file_path)
            wb.close()

            download_url = _get_download_url(filename)

            return BuiltinToolResult(
                success=True,
                result={
                    "filename": filename,
                    "path": file_path,
                    "operation": operation,
                    "sheet_name": sheet_name,
                    "rows_written": len(data),
                    "columns": len(headers) if headers else (len(data[0]) if data else 0),
                    "download_url": download_url,
                    "ttl_hours": WORKSPACE_FILE_TTL_HOURS,
                },
            )

        elif operation == "add_sheet":
            if not os.path.exists(file_path):
                return BuiltinToolResult(success=False, error=f"File not found: {filename}")

            wb = load_workbook(file_path)

            if sheet_name in wb.sheetnames:
                return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' already exists")

            ws = wb.create_sheet(sheet_name)

            if headers:
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            start_row = 2 if headers else 1
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(file_path)
            wb.close()

            download_url = _get_download_url(filename)

            return BuiltinToolResult(
                success=True,
                result={
                    "filename": filename,
                    "path": file_path,
                    "operation": operation,
                    "sheet_name": sheet_name,
                    "rows_written": len(data),
                    "download_url": download_url,
                    "ttl_hours": WORKSPACE_FILE_TTL_HOURS,
                },
            )

        elif operation == "append_rows":
            if not os.path.exists(file_path):
                return BuiltinToolResult(success=False, error=f"File not found: {filename}")

            wb = load_workbook(file_path)

            if sheet_name not in wb.sheetnames:
                return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

            ws = wb[sheet_name]
            next_row = ws.max_row + 1

            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=next_row + row_idx, column=col_idx, value=value)

            wb.save(file_path)
            wb.close()

            download_url = _get_download_url(filename)

            return BuiltinToolResult(
                success=True,
                result={
                    "filename": filename,
                    "path": file_path,
                    "operation": operation,
                    "sheet_name": sheet_name,
                    "rows_appended": len(data),
                    "starting_row": next_row,
                    "download_url": download_url,
                    "ttl_hours": WORKSPACE_FILE_TTL_HOURS,
                },
            )

        elif operation == "update_cell":
            if not os.path.exists(file_path):
                return BuiltinToolResult(success=False, error=f"File not found: {filename}")

            if not cell_updates:
                return BuiltinToolResult(success=False, error="cell_updates required for update_cell operation")

            wb = load_workbook(file_path)

            if sheet_name not in wb.sheetnames:
                return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

            ws = wb[sheet_name]

            updated_cells = []
            for update in cell_updates:
                cell_ref = update.get("cell", "")
                value = update.get("value")
                if cell_ref:
                    ws[cell_ref] = value
                    updated_cells.append(cell_ref)

            wb.save(file_path)
            wb.close()

            download_url = _get_download_url(filename)

            return BuiltinToolResult(
                success=True,
                result={
                    "filename": filename,
                    "path": file_path,
                    "operation": operation,
                    "sheet_name": sheet_name,
                    "cells_updated": updated_cells,
                    "download_url": download_url,
                    "ttl_hours": WORKSPACE_FILE_TTL_HOURS,
                },
            )

        else:
            return BuiltinToolResult(success=False, error=f"Unknown operation: {operation}")

    except Exception as e:
        logger.exception(f"Spreadsheet write failed: {e}")
        return BuiltinToolResult(success=False, error=f"Write failed: {str(e)}")
