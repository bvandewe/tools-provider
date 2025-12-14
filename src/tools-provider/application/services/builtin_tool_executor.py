"""Built-in Tool Executor for local tool execution.

This module handles the execution of built-in utility tools that run
locally within the tools-provider, without proxying to external services.

Security Considerations:
- URL fetching is restricted to HTTP/HTTPS
- Size limits prevent memory exhaustion
- Timeouts prevent hanging
- Math evaluation is sandboxed
"""

import base64
import html
import json
import logging
import math
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import UTC, datetime, timedelta
from typing import Any
from urllib.parse import quote, unquote
from uuid import uuid4

import httpx
from opentelemetry import trace

logger = logging.getLogger(__name__)
tracer = trace.get_tracer(__name__)


# =============================================================================
# Configuration
# =============================================================================

# Maximum content size to fetch (10MB)
MAX_CONTENT_SIZE = 10 * 1024 * 1024

# Request timeout for fetch_url
FETCH_TIMEOUT = 30.0


# =============================================================================
# Result Type
# =============================================================================


@dataclass
class BuiltinToolResult:
    """Result of executing a built-in tool.

    Attributes:
        success: Whether execution succeeded
        result: Result data (on success)
        error: Error message (on failure)
        metadata: Additional execution metadata
    """

    success: bool
    result: Any = None
    error: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class UserContext:
    """User context for scoping built-in tool operations.

    This ensures memory and file operations are isolated per user.

    Attributes:
        user_id: Unique user identifier (from JWT 'sub' claim)
        username: Human-readable username (optional, for logging)
    """

    user_id: str
    username: str | None = None


# =============================================================================
# Main Executor
# =============================================================================


class BuiltinToolExecutor:
    """Executes built-in utility tools locally.

    This executor handles tools registered with the BuiltinSourceAdapter.
    Tools are executed in-process without HTTP proxying.
    """

    # Mapping of tool names to execution functions
    _EXECUTORS: dict[str, Any] = {}

    def __init__(self) -> None:
        """Initialize the executor with tool mappings."""
        self._current_user_context: UserContext | None = None
        self._executors: dict[str, Any] = {
            # Original tools
            "fetch_url": self._execute_fetch_url,
            "get_current_datetime": self._execute_get_current_datetime,
            "calculate": self._execute_calculate,
            "generate_uuid": self._execute_generate_uuid,
            "encode_decode": self._execute_encode_decode,
            "regex_extract": self._execute_regex_extract,
            "json_transform": self._execute_json_transform,
            "text_stats": self._execute_text_stats,
            # Web & Search tools
            "web_search": self._execute_web_search,
            # "browser_navigate": self._execute_browser_navigate,  # Commented out - requires Playwright (~500MB)
            "wikipedia_query": self._execute_wikipedia_query,
            # Code execution
            "execute_python": self._execute_python,
            # File tools
            "file_writer": self._execute_file_writer,
            "file_reader": self._execute_file_reader,
            "spreadsheet_read": self._execute_spreadsheet_read,
            "spreadsheet_write": self._execute_spreadsheet_write,
            # Memory tools
            "memory_store": self._execute_memory_store,
            "memory_retrieve": self._execute_memory_retrieve,
            # Human interaction
            "ask_human": self._execute_ask_human,
        }
        logger.info(f"BuiltinToolExecutor initialized with {len(self._executors)} tools")

    async def execute(
        self,
        tool_name: str,
        arguments: dict[str, Any],
        user_context: UserContext | None = None,
    ) -> BuiltinToolResult:
        """Execute a built-in tool.

        Args:
            tool_name: Name of the tool to execute
            arguments: Arguments for the tool
            user_context: Optional user context for scoping operations (memory, files)

        Returns:
            BuiltinToolResult with execution outcome
        """
        # Store user context for memory operations
        self._current_user_context = user_context

        with tracer.start_as_current_span(f"builtin_tool.{tool_name}") as span:
            span.set_attribute("tool.name", tool_name)
            if user_context:
                span.set_attribute("user.id", user_context.user_id)

            executor = self._executors.get(tool_name)
            if not executor:
                return BuiltinToolResult(
                    success=False,
                    error=f"Unknown built-in tool: {tool_name}",
                )

            try:
                return await executor(arguments)
            except Exception as e:
                logger.exception(f"Built-in tool execution failed: {tool_name}")
                span.set_attribute("tool.error", str(e))
                return BuiltinToolResult(
                    success=False,
                    error=f"Execution failed: {str(e)}",
                )

    def is_builtin_tool(self, tool_name: str) -> bool:
        """Check if a tool name is a built-in tool."""
        return tool_name in self._executors

    # =========================================================================
    # Tool Implementations
    # =========================================================================

    async def _execute_fetch_url(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the fetch_url tool."""
        url = arguments.get("url", "")
        extract_text = arguments.get("extract_text", True)
        save_as_file = arguments.get("save_as_file")  # Optional filename to save binary content

        # Validate URL
        if not url:
            return BuiltinToolResult(success=False, error="URL is required")

        if not url.startswith(("http://", "https://")):
            return BuiltinToolResult(
                success=False,
                error="URL must start with http:// or https://",
            )

        # Validate save_as_file if provided
        if save_as_file:
            if ".." in save_as_file or save_as_file.startswith("/"):
                return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

        logger.info(f"Fetching URL: {url}")

        try:
            async with httpx.AsyncClient(
                timeout=FETCH_TIMEOUT,
                follow_redirects=True,
                max_redirects=5,
            ) as client:
                # First, do a HEAD request to check content type and size
                try:
                    head_response = await client.head(url)
                    content_length = head_response.headers.get("content-length")
                    if content_length and int(content_length) > MAX_CONTENT_SIZE:
                        return BuiltinToolResult(
                            success=False,
                            error=f"Content too large: {int(content_length)} bytes (max: {MAX_CONTENT_SIZE})",
                        )
                except Exception:  # noqa: S110  # nosec B110
                    # HEAD not supported, continue with GET
                    pass

                # Fetch the content
                response = await client.get(url)
                response.raise_for_status()

                content_type = response.headers.get("content-type", "").lower()
                content_length = len(response.content)

                if content_length > MAX_CONTENT_SIZE:
                    return BuiltinToolResult(
                        success=False,
                        error=f"Content too large: {content_length} bytes (max: {MAX_CONTENT_SIZE})",
                    )

                # Handle different content types
                if self._is_text_content(content_type):
                    content = response.text

                    # Extract plain text from HTML if requested
                    if extract_text and "html" in content_type:
                        content = self._extract_text_from_html(content)

                    return BuiltinToolResult(
                        success=True,
                        result=content,
                        metadata={
                            "url": str(response.url),
                            "status_code": response.status_code,
                            "content_length": content_length,
                            "content_type": content_type.split(";")[0].strip(),
                        },
                    )

                elif self._is_json_content(content_type):
                    try:
                        json_content = response.json()
                        return BuiltinToolResult(
                            success=True,
                            result=json_content,
                            metadata={
                                "url": str(response.url),
                                "status_code": response.status_code,
                                "content_length": content_length,
                                "content_type": "application/json",
                            },
                        )
                    except Exception:
                        # If JSON parsing fails, return as text
                        return BuiltinToolResult(
                            success=True,
                            result=response.text,
                            metadata={
                                "url": str(response.url),
                                "status_code": response.status_code,
                                "content_length": content_length,
                                "content_type": content_type.split(";")[0].strip(),
                            },
                        )

                else:
                    # Binary content - always save to workspace (auto-save behavior)
                    # Use provided filename or extract from response/URL
                    filename = save_as_file or self._extract_filename(response, url)

                    # Always save binary files to workspace
                    import os

                    workspace_dir = self._get_workspace_dir()
                    self._cleanup_old_files(workspace_dir)
                    file_path = os.path.join(workspace_dir, filename)

                    # Ensure directory exists
                    os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else workspace_dir, exist_ok=True)

                    with open(file_path, "wb") as f:
                        f.write(response.content)

                    logger.info(f"Auto-saved binary file: {filename} ({content_length} bytes)")

                    return BuiltinToolResult(
                        success=True,
                        result={
                            "message": f"Binary file saved to workspace: {filename}",
                            "filename": filename,
                            "path": file_path,
                            "size_bytes": content_length,
                            "content_type": content_type.split(";")[0].strip(),
                        },
                        metadata={
                            "url": str(response.url),
                            "status_code": response.status_code,
                            "content_length": content_length,
                            "filename": filename,
                            "is_binary": True,
                            "saved_to_workspace": True,
                            "content_type": content_type.split(";")[0].strip(),
                        },
                    )

        except httpx.TimeoutException:
            return BuiltinToolResult(
                success=False,
                error=f"Request timed out after {FETCH_TIMEOUT} seconds",
            )
        except httpx.HTTPStatusError as e:
            return BuiltinToolResult(
                success=False,
                error=f"HTTP error {e.response.status_code}: {e.response.reason_phrase}",
                metadata={"status_code": e.response.status_code},
            )
        except httpx.RequestError as e:
            return BuiltinToolResult(
                success=False,
                error=f"Request failed: {str(e)}",
            )

    async def _execute_get_current_datetime(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the get_current_datetime tool."""
        timezone_name = arguments.get("timezone", "UTC")
        output_format = arguments.get("format", "all")

        try:
            # Get current time in UTC
            now = datetime.now(UTC)

            # Try to convert to requested timezone
            try:
                import zoneinfo

                tz = zoneinfo.ZoneInfo(timezone_name)
                now_tz = now.astimezone(tz)
            except Exception:
                # Fall back to UTC if timezone not found
                now_tz = now
                logger.warning(f"Unknown timezone: {timezone_name}, using UTC")

            # Format based on request
            if output_format == "iso":
                result = now_tz.isoformat()
            elif output_format == "unix":
                result = int(now_tz.timestamp())
            elif output_format == "human":
                result = now_tz.strftime("%A, %B %d, %Y at %I:%M:%S %p %Z")
            else:  # "all"
                result = {
                    "iso": now_tz.isoformat(),
                    "unix": int(now_tz.timestamp()),
                    "human": now_tz.strftime("%A, %B %d, %Y at %I:%M:%S %p %Z"),
                    "date": now_tz.strftime("%Y-%m-%d"),
                    "time": now_tz.strftime("%H:%M:%S"),
                    "timezone": timezone_name,
                }

            return BuiltinToolResult(success=True, result=result)

        except Exception as e:
            return BuiltinToolResult(success=False, error=f"Failed to get datetime: {str(e)}")

    async def _execute_calculate(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the calculate tool with safe math evaluation."""
        expression = arguments.get("expression", "")
        precision = arguments.get("precision", 10)

        if not expression:
            return BuiltinToolResult(success=False, error="Expression is required")

        # Safe math namespace
        safe_dict = {
            "abs": abs,
            "round": round,
            "min": min,
            "max": max,
            "sum": sum,
            "pow": pow,
            "sqrt": math.sqrt,
            "floor": math.floor,
            "ceil": math.ceil,
            "sin": math.sin,
            "cos": math.cos,
            "tan": math.tan,
            "asin": math.asin,
            "acos": math.acos,
            "atan": math.atan,
            "log": math.log,
            "log10": math.log10,
            "log2": math.log2,
            "exp": math.exp,
            "pi": math.pi,
            "e": math.e,
        }

        try:
            # Sanitize expression - only allow safe characters
            if not re.match(r"^[\d\s\+\-\*/\.\(\)\,\w]+$", expression):
                return BuiltinToolResult(
                    success=False,
                    error="Expression contains invalid characters",
                )

            # Evaluate in restricted namespace
            result = eval(expression, {"__builtins__": {}}, safe_dict)  # noqa: S307  # nosec B307 - restricted namespace

            # Round to precision
            if isinstance(result, float):
                result = round(result, precision)

            return BuiltinToolResult(
                success=True,
                result=result,
                metadata={"expression": expression},
            )

        except ZeroDivisionError:
            return BuiltinToolResult(success=False, error="Division by zero")
        except Exception as e:
            return BuiltinToolResult(success=False, error=f"Calculation error: {str(e)}")

    async def _execute_generate_uuid(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the generate_uuid tool."""
        count = min(max(arguments.get("count", 1), 1), 100)
        output_format = arguments.get("format", "standard")

        uuids = []
        for _ in range(count):
            new_uuid = uuid4()
            if output_format == "hex":
                uuids.append(new_uuid.hex)
            elif output_format == "urn":
                uuids.append(new_uuid.urn)
            else:
                uuids.append(str(new_uuid))

        result = uuids[0] if count == 1 else uuids
        return BuiltinToolResult(success=True, result=result)

    async def _execute_encode_decode(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the encode_decode tool."""
        text = arguments.get("text", "")
        encoding = arguments.get("encoding", "")
        operation = arguments.get("operation", "")

        if not text:
            return BuiltinToolResult(success=False, error="Text is required")
        if not encoding:
            return BuiltinToolResult(success=False, error="Encoding is required")
        if operation not in ("encode", "decode"):
            return BuiltinToolResult(success=False, error="Operation must be 'encode' or 'decode'")

        try:
            if encoding == "base64":
                if operation == "encode":
                    result = base64.b64encode(text.encode()).decode()
                else:
                    result = base64.b64decode(text.encode()).decode()

            elif encoding == "url":
                if operation == "encode":
                    result = quote(text, safe="")
                else:
                    result = unquote(text)

            elif encoding == "html":
                if operation == "encode":
                    result = html.escape(text)
                else:
                    result = html.unescape(text)

            elif encoding == "hex":
                if operation == "encode":
                    result = text.encode().hex()
                else:
                    result = bytes.fromhex(text).decode()

            else:
                return BuiltinToolResult(success=False, error=f"Unknown encoding: {encoding}")

            return BuiltinToolResult(success=True, result=result)

        except Exception as e:
            return BuiltinToolResult(success=False, error=f"{operation.title()} error: {str(e)}")

    async def _execute_regex_extract(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the regex_extract tool."""
        text = arguments.get("text", "")
        pattern = arguments.get("pattern", "")
        flags_str = arguments.get("flags", "")
        max_matches = min(arguments.get("max_matches", 100), 1000)

        if not text:
            return BuiltinToolResult(success=False, error="Text is required")
        if not pattern:
            return BuiltinToolResult(success=False, error="Pattern is required")

        # Build flags
        flags = 0
        if "i" in flags_str:
            flags |= re.IGNORECASE
        if "m" in flags_str:
            flags |= re.MULTILINE
        if "s" in flags_str:
            flags |= re.DOTALL

        try:
            compiled = re.compile(pattern, flags)
            matches = []

            for i, match in enumerate(compiled.finditer(text)):
                if i >= max_matches:
                    break

                match_info: dict[str, Any] = {
                    "match": match.group(),
                    "start": match.start(),
                    "end": match.end(),
                }

                # Add named groups if any
                if match.groupdict():
                    match_info["groups"] = match.groupdict()
                elif match.groups():
                    match_info["groups"] = list(match.groups())

                matches.append(match_info)

            return BuiltinToolResult(
                success=True,
                result=matches,
                metadata={"pattern": pattern, "match_count": len(matches)},
            )

        except re.error as e:
            return BuiltinToolResult(success=False, error=f"Invalid regex: {str(e)}")

    async def _execute_json_transform(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the json_transform tool."""
        data = arguments.get("data")
        path = arguments.get("path", "")

        if data is None:
            return BuiltinToolResult(success=False, error="Data is required")
        if not path:
            return BuiltinToolResult(success=False, error="Path is required")

        # Parse JSON string if needed
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError as e:
                return BuiltinToolResult(success=False, error=f"Invalid JSON: {str(e)}")

        try:
            # Simple JSONPath implementation
            result = self._apply_jsonpath(data, path)
            return BuiltinToolResult(success=True, result=result)

        except Exception as e:
            return BuiltinToolResult(success=False, error=f"JSONPath error: {str(e)}")

    async def _execute_text_stats(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the text_stats tool."""
        text = arguments.get("text", "")
        include_freq = arguments.get("include_word_frequency", False)

        if not text:
            return BuiltinToolResult(success=False, error="Text is required")

        # Character counts
        char_count = len(text)
        char_count_no_spaces = len(text.replace(" ", "").replace("\n", "").replace("\t", ""))

        # Word count
        words = re.findall(r"\b\w+\b", text.lower())
        word_count = len(words)

        # Sentence count (simple heuristic)
        sentences = re.split(r"[.!?]+", text)
        sentence_count = len([s for s in sentences if s.strip()])

        # Paragraph count
        paragraphs = re.split(r"\n\s*\n", text)
        paragraph_count = len([p for p in paragraphs if p.strip()])

        # Average word length
        avg_word_length = sum(len(w) for w in words) / word_count if word_count else 0

        # Reading time (average 200 words per minute)
        reading_time_minutes = word_count / 200

        result: dict[str, Any] = {
            "character_count": char_count,
            "character_count_no_spaces": char_count_no_spaces,
            "word_count": word_count,
            "sentence_count": sentence_count,
            "paragraph_count": paragraph_count,
            "average_word_length": round(avg_word_length, 2),
            "reading_time_minutes": round(reading_time_minutes, 1),
        }

        if include_freq:
            word_freq = Counter(words).most_common(10)
            result["top_words"] = [{"word": w, "count": c} for w, c in word_freq]

        return BuiltinToolResult(success=True, result=result)

    # =========================================================================
    # Helper Methods
    # =========================================================================

    def _is_text_content(self, content_type: str) -> bool:
        """Check if content type indicates text content."""
        text_types = [
            "text/",
            "application/xml",
            "application/xhtml",
            "application/csv",
        ]
        return any(t in content_type for t in text_types)

    def _is_json_content(self, content_type: str) -> bool:
        """Check if content type indicates JSON content."""
        return "application/json" in content_type or "+json" in content_type

    def _extract_filename(self, response: httpx.Response, url: str) -> str:
        """Extract filename from response headers or URL."""
        # Try Content-Disposition header
        content_disposition = response.headers.get("content-disposition", "")
        if "filename=" in content_disposition:
            match = re.search(r'filename[*]?=["\']?([^"\';\n]+)', content_disposition)
            if match:
                return match.group(1).strip()

        # Fall back to URL path
        from urllib.parse import urlparse

        path = urlparse(url).path
        if path:
            filename = path.split("/")[-1]
            if filename:
                return filename

        return "unknown"

    def _extract_text_from_html(self, html_content: str) -> str:
        """Extract plain text from HTML content."""
        # Remove script and style elements
        text = re.sub(r"<script[^>]*>.*?</script>", "", html_content, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL | re.IGNORECASE)

        # Remove HTML tags
        text = re.sub(r"<[^>]+>", " ", text)

        # Decode HTML entities (basic)
        text = html.unescape(text)

        # Clean up whitespace
        text = re.sub(r"\s+", " ", text)
        text = text.strip()

        return text

    def _apply_jsonpath(self, data: Any, path: str) -> Any:
        """Apply a simple JSONPath expression to data.

        Supports basic JSONPath:
        - $ (root)
        - .key (object key)
        - [n] (array index)
        - [*] (all array elements)
        - .* (all object values)
        """
        if not path or path == "$":
            return data

        # Remove leading $
        if path.startswith("$"):
            path = path[1:]

        # Remove leading dot
        if path.startswith("."):
            path = path[1:]

        if not path:
            return data

        result = data
        parts = re.findall(r"\.?(\w+|\[\d+\]|\[\*\]|\*)", path)

        for part in parts:
            if result is None:
                return None

            if part == "*" or part == "[*]":
                # Wildcard - collect all elements
                if isinstance(result, list):
                    result = result
                elif isinstance(result, dict):
                    result = list(result.values())
                else:
                    return None

            elif part.startswith("[") and part.endswith("]"):
                # Array index
                try:
                    idx = int(part[1:-1])
                    if isinstance(result, list) and -len(result) <= idx < len(result):
                        result = result[idx]
                    else:
                        return None
                except ValueError:
                    return None

            else:
                # Object key
                if isinstance(result, dict) and part in result:
                    result = result[part]
                elif isinstance(result, list):
                    # Apply to each element
                    result = [item.get(part) if isinstance(item, dict) else None for item in result]
                else:
                    return None

        return result

    # =========================================================================
    # Web & Search Tools
    # =========================================================================

    async def _execute_web_search(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the web_search tool using DuckDuckGo."""
        query = arguments.get("query", "")
        max_results = min(arguments.get("max_results", 5), 10)
        region = arguments.get("region", "wt-wt")

        if not query:
            return BuiltinToolResult(success=False, error="Query is required")

        logger.info(f"Web search: {query}")

        try:
            # Use DuckDuckGo HTML search (no API key required)
            async with httpx.AsyncClient(timeout=FETCH_TIMEOUT) as client:
                response = await client.get(
                    "https://html.duckduckgo.com/html/",
                    params={"q": query, "kl": region},
                    headers={"User-Agent": "Mozilla/5.0 (compatible; MCPToolsProvider/1.0)"},
                )
                response.raise_for_status()

                # Parse results from HTML
                results = self._parse_ddg_results(response.text, max_results)

                return BuiltinToolResult(
                    success=True,
                    result={
                        "query": query,
                        "results": results,
                        "result_count": len(results),
                    },
                )

        except Exception as e:
            logger.exception(f"Web search failed: {e}")
            return BuiltinToolResult(success=False, error=f"Search failed: {str(e)}")

    def _parse_ddg_results(self, html_content: str, max_results: int) -> list[dict]:
        """Parse DuckDuckGo HTML search results."""
        results = []

        # Find result links - DuckDuckGo uses class="result__a"
        link_pattern = r'<a[^>]*class="result__a"[^>]*href="([^"]*)"[^>]*>([^<]*)</a>'
        snippet_pattern = r'<a[^>]*class="result__snippet"[^>]*>([^<]*(?:<[^>]*>[^<]*)*)</a>'

        links = re.findall(link_pattern, html_content)
        snippets = re.findall(snippet_pattern, html_content)

        for i, (url, title) in enumerate(links[:max_results]):
            snippet = snippets[i] if i < len(snippets) else ""
            # Clean snippet HTML
            snippet = re.sub(r"<[^>]*>", "", snippet)
            snippet = html.unescape(snippet).strip()

            # DuckDuckGo wraps URLs - extract real URL
            if "uddg=" in url:
                url_match = re.search(r"uddg=([^&]+)", url)
                if url_match:
                    url = unquote(url_match.group(1))

            results.append(
                {
                    "title": html.unescape(title).strip(),
                    "url": url,
                    "snippet": snippet,
                }
            )

        return results

    async def _execute_browser_navigate(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the browser_navigate tool using Playwright."""
        url = arguments.get("url", "")
        wait_for = arguments.get("wait_for")
        timeout = min(arguments.get("timeout", 30), 60) * 1000  # Convert to ms
        extract_text = arguments.get("extract_text", True)

        if not url:
            return BuiltinToolResult(success=False, error="URL is required")

        if not url.startswith(("http://", "https://")):
            return BuiltinToolResult(success=False, error="URL must start with http:// or https://")

        logger.info(f"Browser navigate: {url}")

        try:
            # Try to import playwright
            try:
                from playwright.async_api import async_playwright
            except ImportError:
                return BuiltinToolResult(
                    success=False,
                    error="Browser navigation requires Playwright. Install with: pip install playwright && playwright install chromium",
                )

            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()

                try:
                    await page.goto(url, timeout=timeout, wait_until="networkidle")

                    if wait_for:
                        await page.wait_for_selector(wait_for, timeout=timeout)

                    if extract_text:
                        content = await page.inner_text("body")
                    else:
                        content = await page.content()

                    # Get page info
                    title = await page.title()
                    final_url = page.url

                finally:
                    await browser.close()

                return BuiltinToolResult(
                    success=True,
                    result=content,
                    metadata={
                        "url": final_url,
                        "title": title,
                        "content_length": len(content),
                    },
                )

        except Exception as e:
            logger.exception(f"Browser navigation failed: {e}")
            return BuiltinToolResult(success=False, error=f"Navigation failed: {str(e)}")

    async def _execute_wikipedia_query(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the wikipedia_query tool."""
        query = arguments.get("query", "")
        language = arguments.get("language", "en")
        sentences = min(arguments.get("sentences", 5), 10)

        if not query:
            return BuiltinToolResult(success=False, error="Query is required")

        logger.info(f"Wikipedia query: {query}")

        try:
            async with httpx.AsyncClient(timeout=FETCH_TIMEOUT) as client:
                # Search for the page
                search_url = f"https://{language}.wikipedia.org/api/rest_v1/page/summary/{quote(query)}"
                response = await client.get(
                    search_url,
                    headers={"User-Agent": "MCPToolsProvider/1.0 (https://github.com/tools-provider)"},
                )

                if response.status_code == 404:
                    # Try searching instead
                    search_api = f"https://{language}.wikipedia.org/w/api.php"
                    search_response = await client.get(
                        search_api,
                        params={
                            "action": "opensearch",
                            "search": query,
                            "limit": 1,
                            "format": "json",
                        },
                    )
                    search_data = search_response.json()

                    if len(search_data) >= 4 and search_data[1]:
                        # Found a match, get that page
                        actual_title = search_data[1][0]
                        response = await client.get(
                            f"https://{language}.wikipedia.org/api/rest_v1/page/summary/{quote(actual_title)}",
                            headers={"User-Agent": "MCPToolsProvider/1.0"},
                        )
                    else:
                        return BuiltinToolResult(
                            success=False,
                            error=f"No Wikipedia article found for: {query}",
                        )

                response.raise_for_status()
                data = response.json()

                # Extract summary (limit sentences if needed)
                extract = data.get("extract", "")
                if sentences < 10 and extract:
                    sentence_list = re.split(r"(?<=[.!?])\s+", extract)
                    extract = " ".join(sentence_list[:sentences])

                return BuiltinToolResult(
                    success=True,
                    result={
                        "title": data.get("title", ""),
                        "summary": extract,
                        "description": data.get("description", ""),
                        "url": data.get("content_urls", {}).get("desktop", {}).get("page", ""),
                        "thumbnail": data.get("thumbnail", {}).get("source", ""),
                    },
                )

        except Exception as e:
            logger.exception(f"Wikipedia query failed: {e}")
            return BuiltinToolResult(success=False, error=f"Wikipedia query failed: {str(e)}")

    # =========================================================================
    # Code Execution Tools
    # =========================================================================

    async def _execute_python(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute Python code in a restricted sandbox."""
        code = arguments.get("code", "")
        timeout = min(arguments.get("timeout", 30), 30)

        if not code:
            return BuiltinToolResult(success=False, error="Code is required")

        logger.info(f"Executing Python code ({len(code)} chars)")

        try:
            # Import RestrictedPython for sandboxing
            try:
                from RestrictedPython import compile_restricted, safe_globals
                from RestrictedPython.Eval import default_guarded_getitem
                from RestrictedPython.Guards import guarded_iter_unpack_sequence, safer_getattr
            except ImportError:
                return BuiltinToolResult(
                    success=False,
                    error="Python execution requires RestrictedPython. Install with: pip install RestrictedPython",
                )

            import io
            import sys
            from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError

            # Compile the code in restricted mode
            byte_code = compile_restricted(code, "<agent_code>", "exec")

            if byte_code.errors:
                return BuiltinToolResult(
                    success=False,
                    error=f"Compilation errors: {byte_code.errors}",
                )

            # Prepare safe globals
            safe_builtins = safe_globals.copy()

            # Add safe modules
            import collections
            import datetime as dt
            import itertools
            import json as json_module
            import math as math_module
            import random
            import re as re_module
            import statistics
            import string

            allowed_modules = {
                "math": math_module,
                "json": json_module,
                "re": re_module,
                "datetime": dt,
                "collections": collections,
                "itertools": itertools,
                "random": random,
                "string": string,
                "statistics": statistics,
            }

            # Capture stdout
            stdout_capture = io.StringIO()
            result_value = {"value": None}

            def execute_code():
                old_stdout = sys.stdout
                try:
                    sys.stdout = stdout_capture
                    local_vars: dict[str, Any] = {}
                    exec_globals = {
                        "__builtins__": safe_builtins,
                        "_getattr_": safer_getattr,
                        "_getitem_": default_guarded_getitem,
                        "_iter_unpack_sequence_": guarded_iter_unpack_sequence,
                        **allowed_modules,
                    }
                    exec(byte_code, exec_globals, local_vars)  # noqa: S102  # nosec B102 - RestrictedPython sandbox
                    # Check for a 'result' variable
                    if "result" in local_vars:
                        result_value["value"] = local_vars["result"]
                finally:
                    sys.stdout = old_stdout

            # Execute with timeout
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(execute_code)
                try:
                    future.result(timeout=timeout)
                except FuturesTimeoutError:
                    return BuiltinToolResult(
                        success=False,
                        error=f"Execution timed out after {timeout} seconds",
                    )

            stdout_output = stdout_capture.getvalue()

            return BuiltinToolResult(
                success=True,
                result={
                    "stdout": stdout_output,
                    "result": result_value["value"],
                },
                metadata={"code_length": len(code)},
            )

        except Exception as e:
            logger.exception(f"Python execution failed: {e}")
            return BuiltinToolResult(success=False, error=f"Execution failed: {str(e)}")

    # =========================================================================
    # File Tools
    # =========================================================================

    async def _execute_file_writer(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the file_writer tool."""
        filename = arguments.get("filename", "")
        content = arguments.get("content", "")
        mode = arguments.get("mode", "overwrite")
        is_binary = arguments.get("is_binary", False)

        if not filename:
            return BuiltinToolResult(success=False, error="Filename is required")
        if not content:
            return BuiltinToolResult(success=False, error="Content is required")

        # Validate filename (security)
        if ".." in filename or filename.startswith("/"):
            return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

        # Validate extension - different allowed sets for text vs binary
        import os

        ext = os.path.splitext(filename)[1].lower()

        # Text file extensions (can be written as text)
        text_extensions = {".txt", ".md", ".json", ".csv", ".py", ".js", ".html", ".css", ".xml", ".yaml", ".yml"}
        # Binary file extensions (require base64 content and is_binary=True)
        binary_extensions = {".xlsx", ".xls", ".pdf", ".png", ".jpg", ".jpeg", ".gif", ".zip", ".docx", ".pptx"}

        allowed_extensions = text_extensions | binary_extensions

        if ext not in allowed_extensions:
            return BuiltinToolResult(
                success=False,
                error=f"File extension '{ext}' not allowed. Allowed: {', '.join(sorted(allowed_extensions))}",
            )

        # Require is_binary=True for binary file types
        if ext in binary_extensions and not is_binary:
            return BuiltinToolResult(
                success=False,
                error=f"Binary extension '{ext}' requires is_binary=True with base64 content. Use spreadsheet_write or fetch_url with save_as_file.",
            )

        # Check size limit (5MB)
        content_size = len(content.encode("utf-8")) if not is_binary else len(content)
        if content_size > 5 * 1024 * 1024:
            return BuiltinToolResult(success=False, error="Content exceeds 5MB limit")

        logger.info(f"Writing file: {filename} ({len(content)} chars, binary={is_binary})")

        try:
            import os

            # Get workspace and run lazy garbage collection
            workspace_dir = self._get_workspace_dir()
            self._cleanup_old_files(workspace_dir)

            file_path = os.path.join(workspace_dir, filename)

            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else workspace_dir, exist_ok=True)

            if is_binary:
                # Binary mode - decode base64 content
                try:
                    binary_content = base64.b64decode(content)
                except Exception as e:
                    return BuiltinToolResult(success=False, error=f"Invalid base64 content: {str(e)}")

                write_mode = "ab" if mode == "append" else "wb"
                with open(file_path, write_mode) as f:
                    f.write(binary_content)

                size_bytes = len(binary_content)
            else:
                # Text mode
                write_mode = "a" if mode == "append" else "w"
                with open(file_path, write_mode, encoding="utf-8") as f:
                    f.write(content)

                size_bytes = len(content.encode("utf-8"))

            # Generate download URL for user access
            download_url = self._get_download_url(filename)

            return BuiltinToolResult(
                success=True,
                result={
                    "filename": filename,
                    "path": file_path,
                    "size_bytes": size_bytes,
                    "mode": mode,
                    "is_binary": is_binary,
                    "download_url": download_url,
                    "ttl_hours": self.WORKSPACE_FILE_TTL_HOURS,
                    "note": f"File available for download at {download_url}. This URL expires in {self.WORKSPACE_FILE_TTL_HOURS} hours.",
                },
            )

        except Exception as e:
            logger.exception(f"File write failed: {e}")
            return BuiltinToolResult(success=False, error=f"Write failed: {str(e)}")

    async def _execute_file_reader(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the file_reader tool."""
        filename = arguments.get("filename", "")
        encoding = arguments.get("encoding", "utf-8")

        if not filename:
            return BuiltinToolResult(success=False, error="Filename is required")

        # Validate filename (security)
        if ".." in filename or filename.startswith("/"):
            return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

        logger.info(f"Reading file: {filename}")

        try:
            import os

            # Get workspace (no cleanup on read - only on write operations)
            workspace_dir = self._get_workspace_dir()
            file_path = os.path.join(workspace_dir, filename)

            if not os.path.exists(file_path):
                return BuiltinToolResult(success=False, error=f"File not found: {filename}")

            # Get file size and extension
            file_size = os.path.getsize(file_path)
            _, ext = os.path.splitext(filename.lower())

            # Check for binary file types that need special handling
            spreadsheet_extensions = {".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}
            document_extensions = {".docx", ".doc", ".pptx", ".ppt", ".odt", ".odp"}
            binary_extensions = {".pdf", ".zip", ".tar", ".gz", ".rar", ".7z", ".exe", ".dll", ".so", ".dylib"}
            image_extensions = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".webp", ".ico"}

            if ext in spreadsheet_extensions:
                return BuiltinToolResult(
                    success=False,
                    error=f"'{filename}' is a spreadsheet file. Use the 'spreadsheet_read' tool instead to read Excel/spreadsheet files.",
                    metadata={"filename": filename, "size_bytes": file_size, "file_type": "spreadsheet"},
                )

            if ext in document_extensions:
                return BuiltinToolResult(
                    success=False,
                    error=f"'{filename}' is a binary document file ({ext}). The file_reader tool only supports text files. "
                    f"File size: {file_size:,} bytes. Consider asking the user to provide the content in a text format.",
                    metadata={"filename": filename, "size_bytes": file_size, "file_type": "document"},
                )

            if ext in binary_extensions:
                return BuiltinToolResult(
                    success=False,
                    error=f"'{filename}' is a binary file ({ext}) that cannot be read as text. File size: {file_size:,} bytes.",
                    metadata={"filename": filename, "size_bytes": file_size, "file_type": "binary"},
                )

            if ext in image_extensions:
                return BuiltinToolResult(
                    success=False,
                    error=f"'{filename}' is an image file ({ext}). The file_reader tool only supports text files. File size: {file_size:,} bytes.",
                    metadata={"filename": filename, "size_bytes": file_size, "file_type": "image"},
                )

            # For text files, read with size limit to prevent huge responses
            max_text_size = 500_000  # 500KB text limit

            with open(file_path, encoding=encoding) as f:
                content = f.read()

            if len(content) > max_text_size:
                # Truncate and indicate there's more
                content = content[:max_text_size]
                return BuiltinToolResult(
                    success=True,
                    result=content,
                    metadata={
                        "filename": filename,
                        "size_bytes": file_size,
                        "truncated": True,
                        "note": f"Content truncated to {max_text_size:,} characters. Total file size: {file_size:,} bytes.",
                    },
                )

            return BuiltinToolResult(
                success=True,
                result=content,
                metadata={
                    "filename": filename,
                    "size_bytes": len(content.encode("utf-8")),
                },
            )

        except UnicodeDecodeError:
            # File is binary but not a known extension - return info about it
            try:
                import os

                file_size = os.path.getsize(file_path)
                return BuiltinToolResult(
                    success=False,
                    error=f"'{filename}' appears to be a binary file that cannot be read as text. File size: {file_size:,} bytes. If this is a spreadsheet, use 'spreadsheet_read' instead.",
                    metadata={"filename": filename, "size_bytes": file_size, "file_type": "binary"},
                )
            except Exception as e:
                return BuiltinToolResult(success=False, error=f"Read failed: {str(e)}")
        except Exception as e:
            logger.exception(f"File read failed: {e}")
            return BuiltinToolResult(success=False, error=f"Read failed: {str(e)}")

    # =========================================================================
    # Spreadsheet Tools
    # =========================================================================

    # Workspace TTL in hours
    WORKSPACE_FILE_TTL_HOURS = 24
    # Maximum cell value length (to prevent huge text cells from breaking SSE)
    MAX_CELL_VALUE_LENGTH = 500

    def _truncate_cell_value(self, value: Any, max_length: int | None = None) -> Any:
        """Truncate cell values that are too long.

        Args:
            value: The cell value to potentially truncate
            max_length: Maximum string length (defaults to MAX_CELL_VALUE_LENGTH)

        Returns:
            Original value or truncated string with indicator
        """
        if value is None:
            return None

        max_len = max_length or self.MAX_CELL_VALUE_LENGTH

        # Only truncate strings
        if isinstance(value, str) and len(value) > max_len:
            return value[:max_len] + f"... [truncated, {len(value)} chars total]"

        return value

    def _get_workspace_dir(self) -> str:
        """Get the agent workspace directory for the current user.

        Uses per-user isolation when user context is available.
        Falls back to a shared 'anonymous' workspace if no user context.
        """
        import os
        import tempfile

        base_dir = os.path.join(tempfile.gettempdir(), "agent_workspace")

        # Use per-user workspace if user context is available
        if self._current_user_context:
            # Sanitize user_id to prevent path traversal
            user_id = self._current_user_context.user_id
            safe_user_id = "".join(c for c in user_id if c.isalnum() or c in "-_")
            if safe_user_id:
                workspace_dir = os.path.join(base_dir, safe_user_id)
            else:
                workspace_dir = os.path.join(base_dir, "anonymous")
        else:
            workspace_dir = os.path.join(base_dir, "anonymous")

        os.makedirs(workspace_dir, exist_ok=True)
        return workspace_dir

    def _get_download_url(self, filename: str) -> str:
        """Generate a download URL for a workspace file.

        Args:
            filename: The filename in the workspace

        Returns:
            Relative URL path for downloading the file
        """
        from urllib.parse import quote

        # Return relative path - the frontend will prepend the base URL
        return f"/api/files/{quote(filename)}"

    def _cleanup_old_files(self, workspace_dir: str, ttl_hours: int = 24) -> int:
        """Remove files older than TTL from workspace. Returns count of removed files."""
        import os
        import time

        removed = 0
        now = time.time()
        ttl_seconds = ttl_hours * 3600

        try:
            for filename in os.listdir(workspace_dir):
                filepath = os.path.join(workspace_dir, filename)
                if os.path.isfile(filepath):
                    file_age = now - os.path.getmtime(filepath)
                    if file_age > ttl_seconds:
                        os.remove(filepath)
                        removed += 1
                        logger.debug(f"Cleaned up old file: {filename}")
        except Exception as e:
            logger.warning(f"Workspace cleanup error: {e}")

        return removed

    async def _execute_spreadsheet_read(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the spreadsheet_read tool."""
        filename = arguments.get("filename", "")
        sheet_name = arguments.get("sheet_name")
        include_stats = arguments.get("include_stats", True)
        max_rows = min(arguments.get("max_rows", 50), 500)  # Reduced default, lower max
        offset = arguments.get("offset", 0)
        columns_filter = arguments.get("columns")

        # Maximum result size in bytes (to prevent SSE issues)
        MAX_RESULT_SIZE = 100_000  # 100KB limit for safe SSE transmission

        if not filename:
            return BuiltinToolResult(success=False, error="Filename is required")

        # Validate filename
        if ".." in filename or filename.startswith("/"):
            return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

        # Validate extension
        if not filename.lower().endswith(".xlsx"):
            return BuiltinToolResult(success=False, error="Only .xlsx files are supported")

        logger.info(f"Reading spreadsheet: {filename}")

        try:
            import json
            import os

            from openpyxl import load_workbook

            # Lazy garbage collection
            workspace_dir = self._get_workspace_dir()
            self._cleanup_old_files(workspace_dir)

            file_path = os.path.join(workspace_dir, filename)

            if not os.path.exists(file_path):
                return BuiltinToolResult(success=False, error=f"File not found: {filename}")

            # Load workbook (read-only for performance)
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # Get dimensions
            all_rows = list(ws.iter_rows(values_only=True))
            total_rows = len(all_rows)

            if total_rows == 0:
                wb.close()
                return BuiltinToolResult(
                    success=True,
                    result={
                        "filename": filename,
                        "sheet_name": sheet_name,
                        "sheets": wb.sheetnames,
                        "total_rows": 0,
                        "headers": [],
                        "data": [],
                    },
                )

            # Extract headers (first row)
            headers = [str(h) if h is not None else f"Column_{i + 1}" for i, h in enumerate(all_rows[0])]

            # Filter columns if specified
            column_indices = None
            if columns_filter:
                column_indices = [i for i, h in enumerate(headers) if h in columns_filter]
                if not column_indices:
                    wb.close()
                    return BuiltinToolResult(success=False, error=f"No matching columns found. Available: {headers}")
                headers = [headers[i] for i in column_indices]

            # Extract data rows with pagination
            data_rows = all_rows[1:]  # Skip header
            total_data_rows = len(data_rows)

            # Apply pagination
            start_idx = offset
            end_idx = min(offset + max_rows, total_data_rows)
            paginated_rows = data_rows[start_idx:end_idx]

            # Convert to list of dicts with size limiting
            data = []
            current_size = 0
            truncated_at_row = None

            for row_idx, row in enumerate(paginated_rows):
                if column_indices:
                    row_data = {headers[j]: self._truncate_cell_value(row[i]) for j, i in enumerate(column_indices)}
                else:
                    row_data = {headers[i]: self._truncate_cell_value(val) for i, val in enumerate(row)}

                # Estimate size of this row
                row_json = json.dumps(row_data, default=str)
                row_size = len(row_json)

                if current_size + row_size > MAX_RESULT_SIZE:
                    truncated_at_row = row_idx
                    break

                data.append(row_data)
                current_size += row_size

            result = {
                "filename": filename,
                "sheet_name": sheet_name,
                "sheets": wb.sheetnames,
                "total_rows": total_data_rows,
                "returned_rows": len(data),
                "offset": offset,
                "headers": headers,
                "data": data,
            }

            # Add truncation warning if needed
            if truncated_at_row is not None:
                result["truncated"] = True
                result["note"] = (
                    f"Results truncated to {len(data)} rows due to size limits. "
                    f"Use 'offset' parameter to paginate through remaining {total_data_rows - offset - len(data)} rows. "
                    f"You can also use 'columns' parameter to select specific columns."
                )

            # Calculate stats if requested
            if include_stats and data:
                stats = {"columns": {}}
                for header in headers:
                    col_values = [row.get(header) for row in data if row.get(header) is not None]
                    col_stats = {"count": len(col_values)}

                    # Check if numeric
                    numeric_values = []
                    for v in col_values:
                        if isinstance(v, int | float):
                            numeric_values.append(v)

                    if numeric_values:
                        col_stats["type"] = "numeric"
                        col_stats["min"] = min(numeric_values)
                        col_stats["max"] = max(numeric_values)
                        col_stats["avg"] = sum(numeric_values) / len(numeric_values)
                    else:
                        col_stats["type"] = "text"
                        col_stats["unique_count"] = len(set(str(v) for v in col_values))

                    stats["columns"][header] = col_stats

                result["stats"] = stats

            wb.close()

            return BuiltinToolResult(
                success=True,
                result=result,
                metadata={
                    "filename": filename,
                    "total_rows": total_data_rows,
                    "returned_rows": len(data),
                },
            )

        except Exception as e:
            logger.exception(f"Spreadsheet read failed: {e}")
            return BuiltinToolResult(success=False, error=f"Read failed: {str(e)}")

    async def _execute_spreadsheet_write(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the spreadsheet_write tool."""
        filename = arguments.get("filename", "")
        operation = arguments.get("operation", "create")
        sheet_name = arguments.get("sheet_name", "Sheet1")
        headers = arguments.get("headers", [])
        data = arguments.get("data", [])
        cell_updates = arguments.get("cell_updates", [])

        if not filename:
            return BuiltinToolResult(success=False, error="Filename is required")

        # Validate filename
        if ".." in filename or filename.startswith("/"):
            return BuiltinToolResult(success=False, error="Invalid filename: path traversal not allowed")

        # Ensure .xlsx extension
        if not filename.lower().endswith(".xlsx"):
            filename = filename + ".xlsx"

        logger.info(f"Writing spreadsheet: {filename} (operation: {operation})")

        try:
            import os

            from openpyxl import Workbook, load_workbook

            # Lazy garbage collection
            workspace_dir = self._get_workspace_dir()
            self._cleanup_old_files(workspace_dir)

            file_path = os.path.join(workspace_dir, filename)

            if operation == "create":
                # Create new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

                # Write headers
                if headers:
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                # Write data
                start_row = 2 if headers else 1
                for row_idx, row_data in enumerate(data, start_row):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                wb.save(file_path)
                wb.close()

                download_url = self._get_download_url(filename)

                return BuiltinToolResult(
                    success=True,
                    result={
                        "filename": filename,
                        "path": file_path,
                        "operation": operation,
                        "sheet_name": sheet_name,
                        "rows_written": len(data),
                        "columns": len(headers) if headers else (len(data[0]) if data else 0),
                        "download_url": download_url,
                        "ttl_hours": self.WORKSPACE_FILE_TTL_HOURS,
                        "note": f"Spreadsheet available at {download_url}. Expires in {self.WORKSPACE_FILE_TTL_HOURS} hours.",
                    },
                )

            elif operation == "add_sheet":
                if not os.path.exists(file_path):
                    return BuiltinToolResult(success=False, error=f"File not found: {filename}")

                wb = load_workbook(file_path)

                if sheet_name in wb.sheetnames:
                    return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' already exists")

                ws = wb.create_sheet(sheet_name)

                # Write headers
                if headers:
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                # Write data
                start_row = 2 if headers else 1
                for row_idx, row_data in enumerate(data, start_row):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                wb.save(file_path)
                wb.close()

                download_url = self._get_download_url(filename)

                return BuiltinToolResult(
                    success=True,
                    result={
                        "filename": filename,
                        "path": file_path,
                        "operation": operation,
                        "sheet_name": sheet_name,
                        "rows_written": len(data),
                        "download_url": download_url,
                        "ttl_hours": self.WORKSPACE_FILE_TTL_HOURS,
                    },
                )

            elif operation == "append_rows":
                if not os.path.exists(file_path):
                    return BuiltinToolResult(success=False, error=f"File not found: {filename}")

                wb = load_workbook(file_path)

                if sheet_name not in wb.sheetnames:
                    return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

                ws = wb[sheet_name]

                # Find next empty row
                next_row = ws.max_row + 1

                # Append data
                for row_idx, row_data in enumerate(data):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=next_row + row_idx, column=col_idx, value=value)

                wb.save(file_path)
                wb.close()

                download_url = self._get_download_url(filename)

                return BuiltinToolResult(
                    success=True,
                    result={
                        "filename": filename,
                        "path": file_path,
                        "operation": operation,
                        "sheet_name": sheet_name,
                        "rows_appended": len(data),
                        "starting_row": next_row,
                        "download_url": download_url,
                        "ttl_hours": self.WORKSPACE_FILE_TTL_HOURS,
                    },
                )

            elif operation == "update_cell":
                if not os.path.exists(file_path):
                    return BuiltinToolResult(success=False, error=f"File not found: {filename}")

                if not cell_updates:
                    return BuiltinToolResult(success=False, error="cell_updates required for update_cell operation")

                wb = load_workbook(file_path)

                if sheet_name not in wb.sheetnames:
                    return BuiltinToolResult(success=False, error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

                ws = wb[sheet_name]

                updated_cells = []
                for update in cell_updates:
                    cell_ref = update.get("cell", "")
                    value = update.get("value")
                    if cell_ref:
                        ws[cell_ref] = value
                        updated_cells.append(cell_ref)

                wb.save(file_path)
                wb.close()

                download_url = self._get_download_url(filename)

                return BuiltinToolResult(
                    success=True,
                    result={
                        "filename": filename,
                        "path": file_path,
                        "operation": operation,
                        "sheet_name": sheet_name,
                        "cells_updated": updated_cells,
                        "download_url": download_url,
                        "ttl_hours": self.WORKSPACE_FILE_TTL_HOURS,
                    },
                )

            else:
                return BuiltinToolResult(success=False, error=f"Unknown operation: {operation}")

        except Exception as e:
            logger.exception(f"Spreadsheet write failed: {e}")
            return BuiltinToolResult(success=False, error=f"Write failed: {str(e)}")

    # =========================================================================
    # Memory Tools
    # =========================================================================

    _redis_memory: Any = None  # Lazy-initialized Redis connection for memory

    async def _get_redis_memory(self) -> Any:
        """Get or create Redis connection for agent memory.

        Returns None if Redis is not configured or unavailable.
        """
        if self._redis_memory is not None:
            return self._redis_memory

        try:
            import redis.asyncio as redis_lib

            from application.settings import Settings

            settings = Settings()
            if not settings.redis_enabled:
                logger.debug("Redis disabled, using file-based memory")
                return None

            self._redis_memory = redis_lib.from_url(
                settings.redis_memory_url,
                decode_responses=True,
            )
            # Test connection
            await self._redis_memory.ping()
            logger.info(f"Redis memory connected: {settings.redis_memory_url}")
            return self._redis_memory

        except Exception as e:
            logger.warning(f"Redis memory unavailable, using file fallback: {e}")
            self._redis_memory = None
            return None

    def _get_memory_key(self, key: str) -> str:
        """Get the full Redis key with prefix and user scoping.

        Memory keys are scoped per user to prevent data leakage.
        Format: {prefix}{user_id}:{key}
        """
        try:
            from application.settings import Settings

            settings = Settings()
            prefix = settings.redis_memory_key_prefix
        except Exception:
            prefix = "agent:memory:"

        # Scope by user if available
        if self._current_user_context and self._current_user_context.user_id:
            return f"{prefix}{self._current_user_context.user_id}:{key}"
        else:
            # Fallback to anonymous namespace (should rarely happen)
            logger.warning("Memory operation without user context - using anonymous namespace")
            return f"{prefix}anonymous:{key}"

    async def _execute_memory_store(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the memory_store tool using Redis with file fallback."""
        key = arguments.get("key", "")
        value = arguments.get("value", "")
        ttl_days = min(arguments.get("ttl_days", 30), 365)

        if not key:
            return BuiltinToolResult(success=False, error="Key is required")
        if not value:
            return BuiltinToolResult(success=False, error="Value is required")

        # Validate key format
        if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", key):
            return BuiltinToolResult(
                success=False,
                error="Key must be alphanumeric with underscores, starting with a letter or underscore",
            )

        logger.info(f"Memory store: {key}")

        # Calculate expiration
        expiry_seconds = ttl_days * 86400 if ttl_days > 0 else None
        expiry_iso = (datetime.now(UTC) + timedelta(days=ttl_days)).isoformat() if ttl_days > 0 else None
        stored_at = datetime.now(UTC).isoformat()

        try:
            redis_client = await self._get_redis_memory()

            if redis_client:
                # Use Redis
                redis_key = self._get_memory_key(key)
                data = json.dumps(
                    {
                        "value": value,
                        "stored_at": stored_at,
                        "expires_at": expiry_iso,
                    }
                )

                if expiry_seconds:
                    await redis_client.setex(redis_key, expiry_seconds, data)
                else:
                    await redis_client.set(redis_key, data)

                return BuiltinToolResult(
                    success=True,
                    result={
                        "key": key,
                        "stored": True,
                        "expires_at": expiry_iso,
                        "storage": "redis",
                    },
                )

            else:
                # File-based fallback
                return await self._memory_store_file(key, value, stored_at, expiry_iso)

        except Exception as e:
            logger.exception(f"Memory store failed: {e}")
            return BuiltinToolResult(success=False, error=f"Store failed: {str(e)}")

    async def _memory_store_file(self, key: str, value: str, stored_at: str, expires_at: str | None) -> BuiltinToolResult:
        """File-based memory storage fallback (user-scoped)."""
        import os
        import tempfile

        # Get user-scoped directory
        user_id = self._current_user_context.user_id if self._current_user_context else "anonymous"
        memory_dir = os.path.join(tempfile.gettempdir(), "agent_memory", user_id)
        os.makedirs(memory_dir, exist_ok=True)
        memory_file = os.path.join(memory_dir, "memory.json")

        # Load existing memory
        memory = {}
        if os.path.exists(memory_file):
            with open(memory_file) as f:
                memory = json.load(f)

        memory[key] = {
            "value": value,
            "stored_at": stored_at,
            "expires_at": expires_at,
        }

        with open(memory_file, "w") as f:
            json.dump(memory, f, indent=2)

        return BuiltinToolResult(
            success=True,
            result={
                "key": key,
                "stored": True,
                "expires_at": expires_at,
                "storage": "file",
            },
        )

    async def _execute_memory_retrieve(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the memory_retrieve tool using Redis with file fallback."""
        key = arguments.get("key")
        default = arguments.get("default")

        logger.info(f"Memory retrieve: {key or 'all keys'}")

        try:
            redis_client = await self._get_redis_memory()

            if redis_client:
                # Use Redis
                return await self._memory_retrieve_redis(redis_client, key, default)
            else:
                # File-based fallback
                return await self._memory_retrieve_file(key, default)

        except Exception as e:
            logger.exception(f"Memory retrieve failed: {e}")
            return BuiltinToolResult(success=False, error=f"Retrieve failed: {str(e)}")

    async def _memory_retrieve_redis(self, redis_client: Any, key: str | None, default: Any) -> BuiltinToolResult:
        """Retrieve from Redis memory (user-scoped)."""
        try:
            from application.settings import Settings

            settings = Settings()
            prefix = settings.redis_memory_key_prefix
        except Exception:
            prefix = "agent:memory:"

        # Build user-scoped prefix for pattern matching
        user_id = self._current_user_context.user_id if self._current_user_context else "anonymous"
        user_prefix = f"{prefix}{user_id}:"

        if key is None:
            # Return all keys for this user only
            pattern = f"{user_prefix}*"
            keys = []
            async for k in redis_client.scan_iter(match=pattern):
                # Strip user prefix to get user-facing key
                keys.append(k.replace(user_prefix, "", 1))

            return BuiltinToolResult(
                success=True,
                result={
                    "keys": keys,
                    "count": len(keys),
                    "storage": "redis",
                },
            )

        # Retrieve specific key
        redis_key = self._get_memory_key(key)
        data = await redis_client.get(redis_key)

        if data:
            parsed = json.loads(data)
            return BuiltinToolResult(
                success=True,
                result={
                    "key": key,
                    "value": parsed["value"],
                    "stored_at": parsed["stored_at"],
                    "storage": "redis",
                },
            )
        elif default is not None:
            return BuiltinToolResult(
                success=True,
                result={
                    "key": key,
                    "value": default,
                    "is_default": True,
                },
            )
        else:
            return BuiltinToolResult(
                success=False,
                error=f"Key not found: {key}",
            )

    async def _memory_retrieve_file(self, key: str | None, default: Any) -> BuiltinToolResult:
        """File-based memory retrieval fallback (user-scoped)."""
        import os
        import tempfile

        # Get user-scoped directory
        user_id = self._current_user_context.user_id if self._current_user_context else "anonymous"
        memory_dir = os.path.join(tempfile.gettempdir(), "agent_memory", user_id)
        memory_file = os.path.join(memory_dir, "memory.json")

        # Load memory
        memory = {}
        if os.path.exists(memory_file):
            with open(memory_file) as f:
                memory = json.load(f)

        # Clean expired entries
        now = datetime.now(UTC)
        memory = {k: v for k, v in memory.items() if v.get("expires_at") is None or datetime.fromisoformat(v["expires_at"]) > now}

        if key is None:
            # Return all keys
            return BuiltinToolResult(
                success=True,
                result={
                    "keys": list(memory.keys()),
                    "count": len(memory),
                    "storage": "file",
                },
            )

        # Retrieve specific key
        if key in memory:
            return BuiltinToolResult(
                success=True,
                result={
                    "key": key,
                    "value": memory[key]["value"],
                    "stored_at": memory[key]["stored_at"],
                    "storage": "file",
                },
            )
        elif default is not None:
            return BuiltinToolResult(
                success=True,
                result={
                    "key": key,
                    "value": default,
                    "is_default": True,
                },
            )
        else:
            return BuiltinToolResult(
                success=False,
                error=f"Key not found: {key}",
            )

    # =========================================================================
    # Human Interaction Tools
    # =========================================================================

    async def _execute_ask_human(self, arguments: dict[str, Any]) -> BuiltinToolResult:
        """Execute the ask_human tool.

        This tool returns a special response that signals the agent should
        pause and wait for human input. The actual pausing logic is handled
        by the agent host, not by this executor.
        """
        question = arguments.get("question", "")
        context = arguments.get("context")
        options = arguments.get("options", [])
        input_type = arguments.get("input_type", "text")

        if not question:
            return BuiltinToolResult(success=False, error="Question is required")

        logger.info(f"Ask human: {question[:50]}...")

        # Return a structured response that the agent host can interpret
        return BuiltinToolResult(
            success=True,
            result={
                "action": "request_human_input",
                "question": question,
                "context": context,
                "options": options,
                "input_type": input_type,
                "awaiting_response": True,
            },
            metadata={
                "requires_human_input": True,
                "input_type": input_type,
            },
        )
